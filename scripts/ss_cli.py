#!/usr/bin/env python3
"""
Shutterstock Contributor CLI

Commands:
  - ss:login     → interactive login that persists session cookies/state
  - ss:quota     → show uploads used in the last 7 days (rolling, cap=500)
  - ss:status    → list recent submissions (read-only snapshot)
  - ss:upload    → uploader flow:
                   * builds queue from Excel (or --files)
                   * enforces 500 / 7d quota (conservative)
                   * opens Upload popup, attaches files
                   * watches ingestion and writes to ledger + Excel
                   * generates CSV metadata and uploads it via the 'Upload CSV' modal

Credentials priority:
1) env: SS_USER / SS_PASS
2) config.json → SHUTTERSTOCK.USERNAME / SHUTTERSTOCK.PASSWORD
3) OS keyring (service 'ssku')

Prereqs:
  pip install -r requirements.txt
  playwright install

Notes:
- Timezone is Europe/Madrid for rolling window (500 per 7 days).
- Scraping is conservative: if DOM/XHR changes, commands attempt to fail gracefully.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
from pathlib import Path
from dataclasses import dataclass
from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
import hashlib
import time

import pytz
from rich.console import Console
from rich.table import Table
from rich import box

try:
    import keyring  # type: ignore
except Exception:
    keyring = None

from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError
from openpyxl import load_workbook

# ────────────────────────────────────────────────────────────────────────────────
# Constants & Paths
# ────────────────────────────────────────────────────────────────────────────────
EUROPE_MADRID = pytz.timezone("Europe/Madrid")
DEFAULT_QUOTA = 500  # images per 7 days
ROLLING_DAYS = 7

HOME = Path.home()
SESSION_DIR = HOME / ".ssku" / "session"
SESSION_DIR.mkdir(parents=True, exist_ok=True)
STORAGE_JSON = SESSION_DIR / "storage.json"

DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)
LEDGER_PATH = DATA_DIR / "ss_quota_ledger.jsonl"
PHOTOS_XLSX = DATA_DIR / "photos.xlsx"

# Config path (override with SSKU_CONFIG env)
DEFAULT_CONFIG_PATH = Path(os.getenv("SSKU_CONFIG", "config.json"))

# Known pages
CONTRIB_ROOT = "https://submit.shutterstock.com/"
PORTFOLIO_URL = "https://submit.shutterstock.com/dashboard"
LOGIN_URL = "https://accounts.shutterstock.com/login"

# “Real browser” fingerprint (adjust UA to your local Chrome if needed)
UA_MAC_CHROME = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0.0.0 Safari/537.36"
)
ACCEPT_LANG = "en-US,en;q=0.9"

console = Console()


# ────────────────────────────────────────────────────────────────────────────────
# Utilities
# ────────────────────────────────────────────────────────────────────────────────
def load_config(path: Optional[Path] = None) -> Dict[str, Any]:
    cfg_path = path or DEFAULT_CONFIG_PATH
    if cfg_path.exists():
        try:
            with cfg_path.open("r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def now_madrid() -> datetime:
    return datetime.now(EUROPE_MADRID)


def parse_iso(dt: str) -> datetime:
    try:
        return datetime.fromisoformat(dt.replace("Z", "+00:00")).astimezone(EUROPE_MADRID)
    except Exception:
        return now_madrid()


def load_storage_state() -> Optional[str]:
    if STORAGE_JSON.exists():
        return str(STORAGE_JSON)
    return None


def save_storage_state(context) -> None:
    context.storage_state(path=str(STORAGE_JSON))


def read_credentials() -> tuple[Optional[str], Optional[str]]:
    # Priority: ENV > config.json > keyring
    user = os.getenv("SS_USER")
    pwd = os.getenv("SS_PASS")

    if not user or not pwd:
        cfg = load_config()
        ss_cfg = cfg.get("SHUTTERSTOCK") or cfg.get("shutterstock") or {}
        user = user or ss_cfg.get("USERNAME") or ss_cfg.get("username")
        pwd = pwd or ss_cfg.get("PASSWORD") or ss_cfg.get("password")

    if (not user or not pwd) and keyring is not None:
        service = "ssku"
        if not user:
            user = keyring.get_password(service, "username")  # may be None
        if user and not pwd:
            pwd = keyring.get_password(service, user)

    return user, pwd


# ────────────────────────────────────────────────────────────────────────────────
# Ledger & Excel helpers
# ────────────────────────────────────────────────────────────────────────────────
def sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def ledger_append(event: Dict[str, Any]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    event = {**event, "timestamp": now_madrid().isoformat()}
    with LEDGER_PATH.open("a", encoding="utf-8") as f:
        f.write(json.dumps(event, ensure_ascii=False) + "\n")


def ledger_entries_within(days: int = ROLLING_DAYS) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    if not LEDGER_PATH.exists():
        return out
    cutoff = now_madrid() - timedelta(days=days)
    with LEDGER_PATH.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                ts_raw = obj.get("timestamp", "")
                if not ts_raw:
                    continue
                ts = parse_iso(str(ts_raw))
                if ts >= cutoff:
                    out.append(obj)
            except Exception:
                continue
    return out


def tally_ledger(days: int = ROLLING_DAYS) -> int:
    return len(ledger_entries_within(days))


def excel_update_status(
    path: Path,
    status: str,
    asset_id: Optional[str] = None,
    uploaded_at: Optional[datetime] = None,
    ingested_at: Optional[datetime] = None,
    xlsx_path: Path = PHOTOS_XLSX,
    sheet: str = "Photos",
) -> None:
    if not xlsx_path.exists():
        return
    wb = load_workbook(filename=str(xlsx_path))
    if sheet not in wb.sheetnames:
        return
    ws = wb[sheet]

    # map columns (case-insensitive)
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:ws.max_column]]

    def col_idx(name: str) -> Optional[int]:
        for i, h in enumerate(headers):
            if h.lower() == name.lower():
                return i + 1
        return None

    # Try to find row by any of these file columns
    file_cols = [col_idx(n) for n in ["abs_path", "rel_path", "relative_path", "filepath", "path", "file"]]
    file_cols = [c for c in file_cols if c]

    # Ensure status columns exist (create if missing at end)
    def ensure_col(name: str) -> int:
        idx = col_idx(name)
        if idx:
            return idx
        ws.cell(row=1, column=ws.max_column + 1, value=name)
        return ws.max_column

    c_status = ensure_col("SS_status")
    c_asset = ensure_col("SS_asset_id")
    c_up_at = ensure_col("SS_uploaded_at")
    c_in_at = ensure_col("SS_ingested_at")

    target_row = None
    for r in range(2, ws.max_row + 1):
        for c in file_cols:
            v = ws.cell(row=r, column=c).value
            if not v:
                continue
            try:
                if Path(str(v)).resolve() == path.resolve():
                    target_row = r
                    break
            except Exception:
                pass
        if target_row:
            break

    if not target_row:
        wb.save(str(xlsx_path))
        return

    ws.cell(row=target_row, column=c_status, value=status)
    if asset_id:
        ws.cell(row=target_row, column=c_asset, value=str(asset_id))
    if uploaded_at:
        ws.cell(row=target_row, column=c_up_at, value=uploaded_at.isoformat())
    if ingested_at:
        ws.cell(row=target_row, column=c_in_at, value=ingested_at.isoformat())

    wb.save(str(xlsx_path))


# ────────────────────────────────────────────────────────────────────────────────
# Portal snapshot & Submissions listing (best-effort)
# ────────────────────────────────────────────────────────────────────────────────
@dataclass
class PortalSnapshot:
    count_last_7d: int
    sample: List[Dict[str, Any]]


def snapshot_portal_uploads(verbose: bool = False) -> PortalSnapshot:
    storage = load_storage_state()
    result = PortalSnapshot(count_last_7d=0, sample=[])
    if storage is None:
        return result

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(storage_state=storage)
        page = context.new_page()
        try:
            page.goto(PORTFOLIO_URL, wait_until="domcontentloaded", timeout=30000)
            timestamps: List[datetime] = []
            for sel in ("time[datetime]", "[data-testid*=item] time[datetime]", "[data-ss-test=date]"):
                for el in page.query_selector_all(sel):
                    dt_attr = el.get_attribute("datetime")
                    if dt_attr:
                        timestamps.append(parse_iso(dt_attr))
            if not timestamps:
                text_content = page.content()
                for token in text_content.split():
                    if token.startswith("202"):
                        try:
                            timestamps.append(parse_iso(token.strip('"\'')))
                        except Exception:
                            pass
            cutoff = now_madrid() - timedelta(days=ROLLING_DAYS)
            recent = [t for t in timestamps if t >= cutoff]
            result.count_last_7d = len(recent)
            result.sample = [{"timestamp": t.isoformat()} for t in sorted(recent, reverse=True)[:5]]
        except PWTimeoutError:
            pass
        except Exception:
            pass
        finally:
            context.storage_state(path=str(STORAGE_JSON))
            browser.close()
    return result


@dataclass
class SubmissionItem:
    asset_id: str
    title: str
    status: str
    submitted_at: Optional[datetime] = None
    reviewed_at: Optional[datetime] = None
    published_at: Optional[datetime] = None
    reason: Optional[str] = None
    url: Optional[str] = None


def list_recent_submissions(limit: int = 50, days: int = 30) -> List[SubmissionItem]:
    storage = load_storage_state()
    if storage is None:
        return []

    items: List[SubmissionItem] = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context(storage_state=storage)
        page = context.new_page()

        captured: List[Dict[str, Any]] = []

        def handle_response(resp):
            try:
                ct = resp.headers.get("content-type", "")
                url = resp.url
                if "application/json" in ct and (
                    "submit" in url or "dashboard" in url or "content" in url or "contributor" in url
                ):
                    data = resp.json()
                    captured.append({"url": url, "data": data})
            except Exception:
                pass

        page.on("response", handle_response)

        try:
            page.goto(PORTFOLIO_URL, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(2000)

            candidate_rows = page.query_selector_all('[data-testid*="item"], [role="row"], article, li')
            for el in candidate_rows[:500]:
                try:
                    status = (el.get_attribute("data-status") or el.get_attribute("aria-label") or "").strip()
                    title_el = (el.query_selector("[data-testid=title]") or el.query_selector("h3, h4, .title"))
                    title = title_el.inner_text().strip() if title_el else ""
                    link_el = el.query_selector("a[href]")
                    href = link_el.get_attribute("href") if link_el else None
                    time_el = el.query_selector("time[datetime]")
                    submitted = parse_iso(time_el.get_attribute("datetime")) if time_el else None
                    asset_id = href.split("/")[-1] if href else ""
                    if title or asset_id:
                        items.append(
                            SubmissionItem(
                                asset_id=asset_id or "",
                                title=title or "",
                                status=status or "",
                                submitted_at=submitted,
                                url=(href if href and href.startswith("http") else ("https://submit.shutterstock.com" + href) if href else None),
                            )
                        )
                except Exception:
                    continue

            cutoff = now_madrid() - timedelta(days=days)

            def maybe_add_from_json(obj: Dict[str, Any]):
                try:
                    aid = str(obj.get("id") or obj.get("asset_id") or obj.get("media_id") or "")
                    title = str(obj.get("title") or obj.get("name") or "")
                    status = str(obj.get("status") or obj.get("review_status") or "")
                    ts_fields = ["submitted_at", "created_at", "created", "submit_time", "uploaded_at"]
                    submitted = None
                    for k in ts_fields:
                        if k in obj and obj[k]:
                            submitted = parse_iso(str(obj[k]))
                            break
                    reviewed = parse_iso(str(obj.get("reviewed_at"))) if obj.get("reviewed_at") else None
                    published = parse_iso(str(obj.get("published_at"))) if obj.get("published_at") else None
                    reason = obj.get("rejection_reason") or obj.get("reason")
                    if submitted and submitted < cutoff:
                        return
                    if aid or title or status:
                        items.append(
                            SubmissionItem(
                                asset_id=aid,
                                title=title,
                                status=status,
                                submitted_at=submitted,
                                reviewed_at=reviewed,
                                published_at=published,
                                reason=str(reason) if reason else None,
                            )
                        )
                except Exception:
                    return

            for cap in captured:
                data = cap.get("data")
                if isinstance(data, dict):
                    for v in data.values():
                        if isinstance(v, list):
                            for obj in v:
                                if isinstance(obj, dict):
                                    maybe_add_from_json(obj)
                if isinstance(data, list):
                    for obj in data:
                        if isinstance(obj, dict):
                            maybe_add_from_json(obj)

        except Exception:
            pass
        finally:
            context.storage_state(path=str(STORAGE_JSON))
            browser.close()

    dedup: Dict[str, SubmissionItem] = {}
    for it in items:
        key = (it.asset_id or it.title or "").strip()
        if key and key not in dedup:
            dedup[key] = it
    out = list(dedup.values())

    out.sort(key=lambda x: x.submitted_at or datetime.min.replace(tzinfo=EUROPE_MADRID), reverse=True)
    return out[:limit]


# ────────────────────────────────────────────────────────────────────────────────
# Excel → upload queue
# ────────────────────────────────────────────────────────────────────────────────
@dataclass
class QueueItem:
    path: Path
    sha256: Optional[str]
    title: Optional[str]
    description: Optional[str]
    keywords: List[str]


def _truthy(val: Any) -> bool:
    return str(val).strip().lower() in {"1", "true", "yes", "y", "on"}


def _get(row: Dict[str, Any], *names: str, default: Any = None):
    for n in names:
        if n in row and row[n] is not None:
            return row[n]
        for k in row.keys():
            if k.lower() == n.lower() and row[k] is not None:
                return row[k]
    return default


def read_photos_excel(xlsx_path: Path = PHOTOS_XLSX, sheet: str = "Photos") -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not xlsx_path.exists():
        console.print(f"[yellow]Excel not found[/]: {xlsx_path}")
        return rows
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if sheet not in wb.sheetnames:
        console.print(f"[yellow]Sheet '{sheet}' not found in {xlsx_path.name}")
        return rows
    ws = wb[sheet]
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))[0:ws.max_column]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        rows.append(row)
    return rows


def build_ss_queue(limit: Optional[int] = None) -> List[QueueItem]:
    data = read_photos_excel()
    out: List[QueueItem] = []
    wanted_status = {"READY_TO_UPLOAD", "FAILED_UPLOAD"}
    for row in data:
        enabled = _truthy(_get(row, "SS_enabled", default="1"))
        status = str(_get(row, "SS_status", default="")).strip().upper()
        if not enabled or (status and status not in wanted_status):
            continue
        path_val = _get(row, "file", "filepath", "path", "relative_path", "rel_path", "abs_path", default=None)
        if not path_val:
            continue
        p_raw = str(path_val)
        p = Path(p_raw)
        if not p.is_absolute():
            p = (Path.cwd() / p).resolve()
        if not p.exists():
            console.print(f"[yellow]Missing file, skipping[/]: {p}")
            continue
        title = _get(row, "SS_title", "title")
        desc = _get(row, "SS_description", "description")
        tags_raw = _get(row, "SS_tags", "keywords") or ""
        if isinstance(tags_raw, str):
            kw = [t.strip() for t in tags_raw.split(",") if t.strip()]
        elif isinstance(tags_raw, list):
            kw = [str(t).strip() for t in tags_raw if str(t).strip()]
        else:
            kw = []
        out.append(QueueItem(path=p, sha256=_get(row, "sha256", "hash"), title=title, description=desc, keywords=kw))
        if limit and len(out) >= limit:
            break
    return out


# ────────────────────────────────────────────────────────────────────────────────
# CSV export for Shutterstock bulk metadata
# ────────────────────────────────────────────────────────────────────────────────
ALLOWED_IMAGE_CATEGORIES = {
    "Abstract","Animals/Wildlife","Arts","Backgrounds/Textures","Beauty/Fashion","Buildings/Landmarks",
    "Business/Finance","Celebrities","Education","Food and drink","Healthcare/Medical","Holidays",
    "Industrial","Interiors","Miscellaneous","Nature","Objects","Parks/Outdoor","People","Religion",
    "Science","Signs/Symbols","Technology","Transportation","Vintage"
}

def _coerce_keywords(val) -> list[str]:
    if not val:
        return []
    if isinstance(val, list):
        toks = [str(x).strip() for x in val if str(x).strip()]
    else:
        toks = [t.strip() for t in str(val).split(",") if t.strip()]
    return toks[:50]

def _coerce_categories(row: dict) -> str:
    cand = []
    for key in ["SS_category1","SS_category2","category1","category2","SS_categories","categories"]:
        v = row.get(key) if key in row else None
        if not v:
            continue
        if isinstance(v, list):
            cand += [str(x).strip() for x in v if str(x).strip()]
        else:
            cand += [t.strip() for t in str(v).split(",") if t.strip()]
    uniq = []
    for c in cand:
        if c not in uniq:
            uniq.append(c)
    return ", ".join(uniq[:2])

def export_shutterstock_csv(plan_items: list, xlsx_rows: list[dict], out_dir: Path = DATA_DIR) -> Path:
    """
    Build a CSV for Shutterstock metadata:
      Columns: Filename, Description, Keywords, Categories
    Filename MUST match uploaded basename (case-sensitive as shown in UI).
    Description is capped at 200 characters.
    """
    out_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = out_dir / f"shutterstock_metadata_{ts}.csv"

    idx_by_abs = {}
    idx_by_name = {}
    for r in xlsx_rows:
        p_val = None
        for k in ["file","filepath","path","relative_path","rel_path","abs_path"]:
            if k in r and r[k]:
                p_val = r[k]
                break
        if not p_val:
            continue
        try:
            p = Path(str(p_val))
            if not p.is_absolute():
                p = (Path.cwd() / p).resolve()
            idx_by_abs[str(p.resolve())] = r
            idx_by_name[p.name] = r
        except Exception:
            continue

    rows = []
    for it in plan_items:
        r = idx_by_abs.get(str(it.path.resolve())) or idx_by_name.get(it.path.name) or {}
        desc = r.get("SS_description") or r.get("description") or r.get("SS_title") or r.get("title") or ""
        desc = str(desc).strip()[:200]
        kw = _coerce_keywords(r.get("SS_tags") or r.get("keywords") or it.keywords)
        cats = _coerce_categories(r)

        rows.append({
            "Filename": it.path.name,
            "Description": desc,
            "Keywords": ", ".join(kw),
            "Categories": cats,
        })

    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Filename","Description","Keywords","Categories"])
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    return csv_path


# ────────────────────────────────────────────────────────────────────────────────
# Playwright helpers: ingestion watch + CSV modal
# ────────────────────────────────────────────────────────────────────────────────
def watch_ingestion(page, files: List[Path], network_bucket: Dict[str, Any], timeout_s: int = 180, debug: bool = False) -> Dict[str, Dict[str, Any]]:
    """
    Waits for each file to reach an 'ingested/processed' state.
    Heuristics:
      1) Global success button: "Go to portfolio" → mark all as INGESTED
      2) Per-file: filename present AND any of {100%, complete, done, processed, uploaded}
      3) Network JSON: try to capture asset_id by matching filename/title
    """
    deadline = time.time() + timeout_s
    results: Dict[str, Dict[str, Any]] = {str(f): {"status": "UPLOADED", "uploaded_at": now_madrid().isoformat()} for f in files}

    success_tokens = {"100%", "complete", "completed", "done", "processed", "ingested", "uploaded", "successful", "success"}

    def mark_all_ingested():
        ts = now_madrid().isoformat()
        for k in results:
            results[k]["status"] = "INGESTED"
            results[k]["ingested_at"] = ts

    def filename_visible(name: str) -> bool:
        try:
            return page.locator(f"text={name}").first.is_visible(timeout=500)
        except Exception:
            return False

    def global_success() -> bool:
        try:
            if page.locator("button:has-text('Go to portfolio')").first.is_visible(timeout=500):
                return True
        except Exception:
            pass
        return False

    while time.time() < deadline:
        try:
            if global_success():
                mark_all_ingested()
                break

            rows = page.locator("div,li,article,tr").all()[:800]
            for p in files:
                key = str(p)
                if results[key].get("status") == "INGESTED":
                    continue
                name = p.name

                name_seen = filename_visible(name)
                if name_seen:
                    for row in rows:
                        try:
                            txt = row.inner_text(timeout=0) or ""
                        except Exception:
                            continue
                        low = txt.lower()
                        if name in txt and any(tok in low for tok in success_tokens):
                            results[key]["status"] = "INGESTED"
                            results[key]["ingested_at"] = now_madrid().isoformat()
                            break

            for key in list(results.keys()):
                if "asset_id" in results[key]:
                    continue
                name = Path(key).name
                for item in network_bucket.get("json", []):
                    try:
                        data = item.get("data")
                        candidates = [data] if isinstance(data, dict) else (data if isinstance(data, list) else [])
                        for obj in candidates:
                            if not isinstance(obj, dict):
                                continue
                            aid = obj.get("id") or obj.get("asset_id") or obj.get("media_id")
                            title = (obj.get("title") or obj.get("name") or "")
                            filename = (obj.get("filename") or obj.get("file_name") or "")
                            if aid and (name in str(filename) or name in str(title)):
                                results[key]["asset_id"] = str(aid)
                                break
                    except Exception:
                        continue

            if all(results[str(p)].get("status") == "INGESTED" for p in files):
                break

            time.sleep(0.8)
        except Exception:
            time.sleep(0.8)

    if debug:
        try:
            dbg_path = DATA_DIR / "ss_last_network.json"
            with dbg_path.open("w", encoding="utf-8") as f:
                json.dump(network_bucket, f, ensure_ascii=False, indent=2)
            console.print(f"[dim]Saved network debug to {dbg_path}")
        except Exception:
            pass

    return results


def upload_metadata_csv(page, csv_path: Path, timeout_s: int = 90) -> bool:
    """
    Opens the 'Upload Metadata from CSV' modal and uploads the given CSV.
    Returns True if we see success-ish cues (toast, message), False otherwise.
    """
    try:
        if page.locator("[data-testid='csv-upload']").first.count():
            page.locator("[data-testid='csv-upload']").first.click(timeout=4000)
        else:
            page.locator("button:has-text('Upload CSV')").first.click(timeout=4000)
    except Exception:
        try:
            page.locator("button[aria-label*=more], [data-testid*=more]").first.click(timeout=2000)
            page.locator("text=Upload CSV").first.click(timeout=2000)
        except Exception:
            return False

    page.wait_for_timeout(600)

    file_inputs = [
        "input[type=file]",
        "input[type='file']",
        "form [type=file]",
        "[data-testid='file-input'] input[type=file]"
    ]
    input_sel = None
    for sel in file_inputs:
        try:
            el = page.locator(sel).first
            if el and el.count():
                input_sel = sel
                break
        except Exception:
            continue
    if not input_sel:
        return False

    try:
        page.set_input_files(input_sel, str(csv_path))
    except Exception:
        try:
            page.locator(input_sel).scroll_into_view_if_needed(timeout=1000)
            page.set_input_files(input_sel, str(csv_path))
        except Exception:
            return False

    end = time.time() + timeout_s
    ok_tokens = {"processing", "processed", "success", "uploaded", "applied", "complete"}
    while time.time() < end:
        try:
            toast = page.locator("div[role=alert], [data-testid=toast], [class*=toast]").first
            if toast.count():
                txt = (toast.inner_text(timeout=0) or "").lower()
                if any(t in txt for t in ok_tokens):
                    return True
        except Exception:
            pass
        try:
            if page.locator("text=Your CSV has been uploaded").first.is_visible(timeout=500):
                return True
        except Exception:
            pass
        time.sleep(0.7)
    return False


# ────────────────────────────────────────────────────────────────────────────────
# Commands
# ────────────────────────────────────────────────────────────────────────────────
def cmd_login(args: argparse.Namespace) -> int:
    with sync_playwright() as pw:
        launch_kwargs = {
            "headless": False,
            "slow_mo": 50,
            "args": ["--disable-blink-features=AutomationControlled"],
        }
        try:
            browser = pw.chromium.launch(channel="chrome", **launch_kwargs)
        except Exception:
            browser = pw.chromium.launch(**launch_kwargs)

        context = browser.new_context(
            storage_state=load_storage_state(),
            user_agent=UA_MAC_CHROME,
            locale="en-US",
            extra_http_headers={"Accept-Language": ACCEPT_LANG},
        )
        page = context.new_page()
        console.rule("[bold]Shutterstock Contributor Login")
        page.goto(CONTRIB_ROOT, wait_until="domcontentloaded")

        try:
            page.wait_for_selector("a[href*='dashboard']", timeout=4000)
            console.print("[green]Already logged in. Session refreshed.")
            save_storage_state(context)
            browser.close()
            return 0
        except PWTimeoutError:
            pass

        page.goto(LOGIN_URL, wait_until="domcontentloaded")

        user, pwd = read_credentials()
        if user and pwd:
            try:
                page.fill("input[type=email], input[name=email]", user)
                page.fill("input[type=password]", pwd)
                page.click("button[type=submit], button[data-testid='login-submit']")
            except Exception:
                pass

        console.print("If prompted, complete any MFA challenge in the opened window…")
        try:
            page.wait_for_url(lambda url: "submit.shutterstock.com" in url or "account" not in url, timeout=180000)
        except PWTimeoutError:
            console.print("[red]Login did not complete within 3 minutes.]")
            save_storage_state(context)
            browser.close()
            return 2

        try:
            page.goto(CONTRIB_ROOT, wait_until="domcontentloaded")
            page.wait_for_selector("a[href*='dashboard']", timeout=15000)
            console.print("[green]Login successful. Session stored.]")
            save_storage_state(context)
            browser.close()
            return 0
        except PWTimeoutError:
            console.print("[yellow]Unable to verify dashboard element. Session saved anyway.]")
            save_storage_state(context)
            browser.close()
            return 1


def cmd_quota(args: argparse.Namespace) -> int:
    ledger_count = tally_ledger(ROLLING_DAYS)
    snap = snapshot_portal_uploads(verbose=args.verbose)
    used = min(ledger_count, snap.count_last_7d) if snap.count_last_7d else ledger_count
    remaining = max(0, DEFAULT_QUOTA - used)

    console.rule("[bold]Shutterstock 7-day Upload Quota")
    table = Table(box=box.SIMPLE_HEAVY)
    table.add_column("Window", justify="left")
    table.add_column("Used", justify="right")
    table.add_column("Remaining", justify="right")
    start = (now_madrid() - timedelta(days=ROLLING_DAYS)).strftime("%Y-%m-%d %H:%M %Z")
    end = now_madrid().strftime("%Y-%m-%d %H:%M %Z")
    table.add_row(f"{start} → {end}", str(used), str(remaining))
    console.print(table)

    if args.verbose:
        vt = Table(title="Details", box=box.MINIMAL_DOUBLE_HEAD)
        vt.add_column("Source")
        vt.add_column("Count")
        vt.add_row("Local ledger (last 7d)", str(ledger_count))
        vt.add_row("Portal snapshot (last 7d)", str(snap.count_last_7d))
        console.print(vt)
        if snap.sample:
            st = Table(title="Portal recent sample", box=box.SIMPLE)
            st.add_column("timestamp (Europe/Madrid)")
            for item in snap.sample:
                st.add_row(item.get("timestamp", ""))
            console.print(st)

    if remaining <= 0:
        console.print("[red]Quota appears exhausted. Delay new uploads to avoid rejection.]")
    elif remaining < 25:
        console.print("[yellow]Quota nearly exhausted—consider pausing bulk jobs.]")
    else:
        console.print("[green]Quota headroom looks OK.]")
    return 0


def cmd_status(args: argparse.Namespace) -> int:
    data = list_recent_submissions(limit=args.limit, days=args.days)
    table = Table(title="Shutterstock — Recent Submissions", box=box.SIMPLE_HEAVY)
    table.add_column("ID", justify="left", no_wrap=True)
    table.add_column("Status", justify="left")
    table.add_column("Submitted", justify="left")
    table.add_column("Reviewed", justify="left")
    table.add_column("Published", justify="left")
    table.add_column("Title", justify="left")

    def fmt(dt: Optional[datetime]):
        return dt.astimezone(EUROPE_MADRID).strftime("%Y-%m-%d %H:%M") if dt else ""

    for it in data:
        short_id = (it.asset_id[:10] + ("…" if len(it.asset_id) > 10 else "")) if it.asset_id else ""
        title = (it.title[:60] + "…") if len(it.title) > 60 else it.title
        table.add_row(short_id, it.status or "", fmt(it.submitted_at), fmt(it.reviewed_at), fmt(it.published_at), title)

    console.print(table)

    if args.verbose:
        for it in data:
            if it.reason:
                console.print(f"[yellow]Note[/]: {it.asset_id or it.title}: {it.reason}")
            if it.url:
                console.print(f"[blue]URL[/]: {it.url}")

    return 0


# ────────────────────────────────────────────────────────────────────────────────
# Upload: dashboard → Upload popup → attach files → watch ingestion → CSV upload
# ────────────────────────────────────────────────────────────────────────────────
def cmd_upload(args: argparse.Namespace) -> int:
    # 1) Build queue from Excel (or fallback to --files)
    queue: List[QueueItem]
    if args.files:
        queue = []
        for f in args.files:
            p = Path(f)
            if not p.exists():
                console.print(f"[yellow]Missing file, skipping[/]: {p}")
                continue
            queue.append(QueueItem(path=p, sha256=None, title=None, description=None, keywords=[]))
    else:
        queue = build_ss_queue(limit=args.limit)

    if not queue:
        console.print("[yellow]No candidates found for upload.")
        return 0

    # 2) Enforce 500 / 7-day quota (conservative)
    ledger_count = tally_ledger(ROLLING_DAYS)
    snap = snapshot_portal_uploads(verbose=False)
    used = min(ledger_count, snap.count_last_7d) if snap.count_last_7d else ledger_count
    remaining = max(0, DEFAULT_QUOTA - used)

    headroom = remaining
    if args.limit is not None:
        headroom = min(headroom, args.limit)

    plan = queue[: headroom]

    console.rule("[bold]Upload plan")
    t = Table(box=box.SIMPLE)
    t.add_column("#", justify="right")
    t.add_column("File")
    t.add_column("Title")
    t.add_column("Keywords")
    for i, it in enumerate(plan, 1):
        t.add_row(str(i), str(it.path), (it.title or ""), str(len(it.keywords)))
    console.print(t)

    console.print(f"Remaining quota estimate: {remaining}")
    console.print(f"This run will attempt: {len(plan)} file(s)")

    if args.dry_run:
        console.print("[green]Dry-run: no browser automation performed.]")
        return 0

    # 3) Open dashboard → click Upload → wait modal → set input files
    storage = load_storage_state()
    if storage is None:
        console.print("[red]No session found. Run ss:login first.]")
        return 2

    candidate_urls = [
        "https://submit.shutterstock.com/dashboard",
        "https://submit.shutterstock.com/dashboard/submit-content",
    ]

    header_upload_btns = [
        "[data-testid='desktop-upload-button']",
        "button:has-text('Upload')",
        "a:has-text('Upload')",
    ]
    modal_upload_btns = [
        "[data-testid='uploadButton']",
        "button:has-text('Upload assets')",
        "[data-testid='upload-assets-button']",
    ]

    file_selectors = [
        "input[type=file]",
        "input[type='file']",
        "[data-testid='file-input'] input[type=file]",
        "input[name=file]",
    ]

    with sync_playwright() as pw:
        launch_kwargs = {
            "headless": not args.headful,
            "args": ["--disable-blink-features=AutomationControlled"],
        }
        try:
            browser = pw.chromium.launch(channel="chrome", **launch_kwargs)
        except Exception:
            browser = pw.chromium.launch(**launch_kwargs)

        context = browser.new_context(
            storage_state=storage,
            user_agent=UA_MAC_CHROME,
            locale="en-US",
            extra_http_headers={"Accept-Language": ACCEPT_LANG},
        )
        page = context.new_page()

        # capture interesting JSON responses (ids, filenames, statuses)
        network_bucket: Dict[str, Any] = {"json": []}

        def on_response(resp):
            try:
                ct = resp.headers.get("content-type", "")
                if "application/json" in ct:
                    url = resp.url
                    if any(k in url for k in ["upload", "media", "assets", "ingest", "submit", "content"]):
                        data = resp.json()
                        network_bucket["json"].append({"url": url, "data": data})
            except Exception:
                pass

        page.on("response", on_response)

        def goto_resilient(url: str, timeout: int = 45000) -> str:
            try:
                page.goto(url, wait_until="domcontentloaded", timeout=timeout)
            except Exception:
                pass
            try:
                page.wait_for_load_state("domcontentloaded", timeout=timeout)
            except Exception:
                pass
            return page.url

        def ensure_logged_in() -> bool:
            try:
                current = page.url
            except Exception:
                current = ""
            if ("contributor-accounts.shutterstock.com/login" not in current and
                "accounts.shutterstock.com/login" not in current):
                return True

            user, pwd = read_credentials()
            if not (user and pwd):
                console.print("[red]We hit login but no credentials available. Run ss:login or set SS_USER/SS_PASS.]")
                return False

            console.print("[yellow]Session needs re-auth. Attempting automatic login…]")
            try:
                page.wait_for_selector("input[type=email], input[name=email], input[name=username]", timeout=20000)
                if page.query_selector("input[type=email]"):
                    page.fill("input[type=email]", user)
                elif page.query_selector("input[name=email]"):
                    page.fill("input[name=email]", user)
                else:
                    page.fill("input[name=username]", user)

                if page.query_selector("input[type=password]"):
                    page.fill("input[type=password]", pwd)
                else:
                    page.fill("input[name=password]", pwd)

                if page.query_selector("button[type=submit]"):
                    page.click("button[type=submit]")
                else:
                    page.click("button[data-testid='login-submit']")
            except Exception:
                console.print("[red]Could not find login form elements automatically.]")
                return False

            try:
                page.wait_for_url(lambda u: "submit.shutterstock.com" in u, timeout=120000)
            except Exception:
                final = goto_resilient("https://submit.shutterstock.com/dashboard", timeout=60000)
                if "submit.shutterstock.com" not in final:
                    console.print("[red]Auto-login didn’t complete (MFA may be required). Try `ss:login` and rerun.]")
                    return False
            return True

        def find_file_input():
            for sel in file_selectors:
                try:
                    el = page.query_selector(sel)
                    if el:
                        return sel
                except Exception:
                    pass
            return None

        def click_first(locators: List[str]) -> bool:
            for css in locators:
                try:
                    loc = page.locator(css).first
                    if loc:
                        loc.click(timeout=4000)
                        return True
                except Exception:
                    continue
            return False

        try:
            for url in candidate_urls:
                goto_resilient(url, timeout=45000)
                page.wait_for_timeout(800)

            if not ensure_logged_in():
                return 2

            goto_resilient("https://submit.shutterstock.com/dashboard", timeout=60000)
            page.wait_for_timeout(800)

            opened = click_first(header_upload_btns)
            if not opened:
                if not ensure_logged_in():
                    return 2

            clicked_modal = click_first(modal_upload_btns)
            if not clicked_modal and not opened:
                console.print("[yellow]Neither header nor modal upload controls were clickable yet; continuing scan…]")

            page.wait_for_timeout(1000)

            sel = find_file_input()
            if not sel:
                for frame in page.frames:
                    try:
                        el = frame.query_selector("input[type=file]")
                        if el:
                            sel = "iframe input[type=file]"
                            break
                    except Exception:
                        pass

            if not sel:
                console.print("[red]No file input found after opening upload UI.]")
                console.print(f"Current URL: {page.url}")
                return 3

            file_paths = [str(it.path) for it in plan]
            try:
                page.set_input_files(sel, file_paths)
            except Exception:
                try:
                    page.locator(sel).scroll_into_view_if_needed(timeout=2000)
                    page.set_input_files(sel, file_paths)
                except Exception as e:
                    console.print(f"[red]Failed to attach files: {e}]")
                    return 4

            console.print(f"[green]Attached {len(file_paths)} file(s) to uploader.]")
            console.print("Watching ingestion (up to 3 minutes)…")

            files_paths = [Path(fp) for fp in file_paths]
            results = watch_ingestion(page, files_paths, network_bucket, timeout_s=180, debug=args.debug)

            # Persist: ledger + Excel
            for p in files_paths:
                rec = results[str(p)]
                uploaded_at = parse_iso(rec.get("uploaded_at")) if rec.get("uploaded_at") else now_madrid()
                ingested_at = parse_iso(rec.get("ingested_at")) if rec.get("ingested_at") else None
                asset_id = rec.get("asset_id")
                file_sha = sha256_file(p)
                ledger_append({"platform": "shutterstock", "event": "upload_started", "file": str(p), "sha256": file_sha})
                if rec.get("status") == "INGESTED":
                    ledger_append({"platform": "shutterstock", "event": "ingested", "file": str(p), "sha256": file_sha, "asset_id": asset_id})
                excel_update_status(p, status=rec.get("status") or "UPLOADED", asset_id=asset_id, uploaded_at=uploaded_at, ingested_at=ingested_at)

            # Build and upload the CSV metadata for all files ingested in this run
            ingested_paths = [p for p in files_paths if (results[str(p)].get("status") == "INGESTED")]
            if ingested_paths:
                plan_index = {str(it.path.resolve()): it for it in plan}
                excel_rows = read_photos_excel()
                mini_plan = []
                for p in ingested_paths:
                    it = plan_index.get(str(p.resolve()))
                    if it:
                        mini_plan.append(it)
                csv_out = export_shutterstock_csv(mini_plan, excel_rows, DATA_DIR)
                console.print(f"[blue]CSV prepared[/]: {csv_out.name}")

                # We need to navigate to the Not submitted page to see Upload CSV
                goto_resilient("https://submit.shutterstock.com/portfolio/not_submitted/photo", timeout=45000)
                page.wait_for_timeout(1000)

                ok_csv = upload_metadata_csv(page, csv_out, timeout_s=120)
                if ok_csv:
                    console.print("[green]CSV uploaded to Shutterstock. Waiting briefly for metadata to apply…]")
                    page.wait_for_timeout(1500)
                    for p in ingested_paths:
                        ledger_append({
                            "platform": "shutterstock",
                            "event": "csv_uploaded",
                            "file": str(p),
                            "sha256": sha256_file(p),
                            "csv": str(csv_out.name),
                        })
                else:
                    console.print("[yellow]CSV upload did not confirm success. Check the modal or try again.]")

            # Human summary
            table = Table(title="Ingestion summary", box=box.SIMPLE_HEAVY)
            table.add_column("File")
            table.add_column("Status")
            table.add_column("Asset ID")
            for p in files_paths:
                rec = results[str(p)]
                table.add_row(p.name, rec.get("status", ""), str(rec.get("asset_id") or ""))
            console.print(table)

            if any(results[str(p)].get("status") != "INGESTED" for p in files_paths):
                console.print("[yellow]Some files did not reach an 'ingested' state within the timeout. They may complete shortly.]")
            else:
                console.print("[green]All files show as ingested. Quota ledger and Excel were updated.]")
        finally:
            context.storage_state(path=str(STORAGE_JSON))
            browser.close()

    return 0


# ────────────────────────────────────────────────────────────────────────────────
# Main CLI
# ────────────────────────────────────────────────────────────────────────────────
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Shutterstock Contributor CLI")
    sub = p.add_subparsers(dest="cmd", required=True)

    sp_login = sub.add_parser("ss:login", help="Interactive login; persist session")
    sp_login.set_defaults(func=cmd_login)

    sp_quota = sub.add_parser("ss:quota", help="Show last-7d uploads vs 500 limit")
    sp_quota.add_argument("--verbose", action="store_true", help="Show details")
    sp_quota.set_defaults(func=cmd_quota)

    sp_status = sub.add_parser("ss:status", help="List recent submissions")
    sp_status.add_argument("--limit", type=int, default=50, help="Max items to show")
    sp_status.add_argument("--days", type=int, default=30, help="Look back this many days")
    sp_status.add_argument("--verbose", action="store_true", help="Show reasons/URLs when available")
    sp_status.set_defaults(func=cmd_status)

    sp_upload = sub.add_parser("ss:upload", help="Upload: plan + popup + attach + watch ingestion + CSV upload")
    sp_upload.add_argument("--limit", type=int, default=None, help="Max files from Excel to consider")
    sp_upload.add_argument("--dry-run", action="store_true", help="Print plan only, no browser work")
    sp_upload.add_argument("--headful", action="store_true", help="Show browser window instead of headless")
    sp_upload.add_argument("--files", nargs="*", default=None, help="Explicit file paths to upload (bypass Excel)")
    sp_upload.add_argument("--debug", action="store_true", help="Dump last network JSON for troubleshooting")
    sp_upload.set_defaults(func=cmd_upload)

    return p


def main(argv: Optional[List[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
