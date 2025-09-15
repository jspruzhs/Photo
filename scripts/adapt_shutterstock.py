import os
import json
import logging
import signal
import atexit
import time
from logging.handlers import RotatingFileHandler
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, Optional, Tuple, List

from openpyxl import load_workbook

import argparse

# ---------- Paths / Constants ----------
EXCEL_PATH = Path("data/photos.xlsx")
PHOTOS_SHEET = "Photos"
PLATFORMS_SHEET = "Platforms"

LOG_DIR = Path("logs")
LOG_FILE = LOG_DIR / "adapt_ss.log"
LOCK_FILE = Path(".adapt_ss.lock")
LOCK_STALE_AFTER = timedelta(hours=2)

GPT_MODEL = "gpt-4o-mini"
TITLE_MAX_CHARS = 70
DEFAULT_DESC_MAX = 200
DEFAULT_MAX_TAGS = 50
MIN_TAGS = 7
PER_ITEM_SLEEP = 0.2  # gentle pacing

# Columns we need on Photos sheet
REQUIRED_COLS = [
    "base_title","base_description","base_tags",
    "SS_enabled","SS_status","SS_title","SS_description","SS_tags",
]

# ---------- Logging ----------
def setup_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("adapt_ss")
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        fh = RotatingFileHandler(LOG_FILE, maxBytes=1_000_000, backupCount=3, encoding="utf-8")
        fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
        fh.setFormatter(fmt)
        logger.addHandler(fh)
        ch = logging.StreamHandler()
        ch.setFormatter(fmt)
        logger.addHandler(ch)
    return logger

log = setup_logging()

# ---------- Locking ----------
_lock_acquired = False
def _remove_lock():
    global _lock_acquired
    try:
        if _lock_acquired and LOCK_FILE.exists():
            LOCK_FILE.unlink()
            log.info("Lock released")
    except Exception as e:
        log.warning(f"Failed to remove lock file: {e}")

def acquire_lock():
    global _lock_acquired
    if LOCK_FILE.exists():
        try:
            mtime = datetime.fromtimestamp(LOCK_FILE.stat().st_mtime)
            age = datetime.now() - mtime
            if age > LOCK_STALE_AFTER:
                log.warning(f"Existing lock is stale (age {age}). Overriding.")
                LOCK_FILE.unlink(missing_ok=True)
            else:
                pid_txt = ""
                try: pid_txt = LOCK_FILE.read_text().strip()
                except Exception: pass
                log.error(f"Another run appears active (pid {pid_txt}, age {age}). Exiting.")
                raise SystemExit(1)
        except FileNotFoundError:
            pass
    try:
        fd = os.open(str(LOCK_FILE), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w") as f:
            f.write(str(os.getpid()))
        _lock_acquired = True
        log.info(f"Lock acquired (pid {os.getpid()})")
    except FileExistsError:
        log.error("Lock already exists, exiting.")
        raise SystemExit(1)

    atexit.register(_remove_lock)
    def _sig(signum, frame):
        log.info(f"Received signal {signum}, cleaning up…")
        _remove_lock()
        raise SystemExit(130)
    for s in (signal.SIGINT, signal.SIGTERM):
        try: signal.signal(s, _sig)
        except Exception: pass

# ---------- Helpers ----------
def load_api_key() -> Optional[str]:
    cfg = Path("config.json")
    if cfg.exists():
        try:
            with open(cfg) as f:
                data = json.load(f)
                if data.get("OPENAI_API_KEY"):
                    return data["OPENAI_API_KEY"]
        except Exception as e:
            log.warning(f"config.json read failed: {e}")
    return os.getenv("OPENAI_API_KEY")

def open_workbook():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"{EXCEL_PATH} not found")
    wb = load_workbook(EXCEL_PATH)
    if PHOTOS_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{PHOTOS_SHEET}' not found")
    ws = wb[PHOTOS_SHEET]
    headers = [c.value for c in ws[1]]

    # Ensure Shutterstock columns exist (append if missing)
    new_cols = []
    for h in REQUIRED_COLS:
        if h not in headers:
            headers.append(h)
            ws.cell(row=1, column=len(headers), value=h)
            new_cols.append(h)
    if new_cols:
        log.info(f"Added missing columns: {', '.join(new_cols)}")
        wb.save(EXCEL_PATH)

    col = {h:i+1 for i,h in enumerate(headers)}
    return wb, ws, headers, col

def parse_bool(val) -> bool:
    if val is True: return True
    if val is False: return False
    if val is None: return False
    s = str(val).strip().lower()
    return s in {"1","true","yes","y","on"}

def read_platform_rules(wb) -> Tuple[int, int]:
    """
    Returns (max_keywords, desc_max_chars) for SS from Platforms sheet if present,
    otherwise defaults (50, 200).
    """
    try:
        ws = wb[PLATFORMS_SHEET]
    except KeyError:
        return (DEFAULT_MAX_TAGS, DEFAULT_DESC_MAX)

    # Expected header row: code | name | max_keywords | description_max_chars | title_required | notes
    hdr = [c.value for c in ws[1]]
    idx = {n:i for i,n in enumerate(hdr)}
    for r in range(2, ws.max_row+1):
        code = str(ws.cell(row=r, column=idx.get("code",0)+1).value or "").strip()
        if code == "SS":
            try:
                mk = int(ws.cell(row=r, column=idx.get("max_keywords",0)+1).value or DEFAULT_MAX_TAGS)
            except Exception:
                mk = DEFAULT_MAX_TAGS
            try:
                dm = int(ws.cell(row=r, column=idx.get("description_max_chars",0)+1).value or DEFAULT_DESC_MAX)
            except Exception:
                dm = DEFAULT_DESC_MAX
            return (mk, dm)
    return (DEFAULT_MAX_TAGS, DEFAULT_DESC_MAX)

def small_qa(title: str, desc: str, tags: List[str], max_tags: int, desc_max: int) -> Tuple[bool, str]:
    if not title or not desc or not tags:
        return False, "Missing one of title/description/tags"
    if len(title) > TITLE_MAX_CHARS:
        return False, "Title too long"
    if len(desc) > desc_max:
        return False, "Description too long"
    if len(tags) < MIN_TAGS:
        return False, f"Too few tags (<{MIN_TAGS})"
    if len(tags) > max_tags:
        return False, f"Too many tags (>{max_tags})"
    # basic dup check
    tl = [t.strip().lower() for t in tags if t.strip()]
    if len(set(tl)) != len(tl):
        return False, "Duplicate tags"
    return True, "ok"

# ---------- GPT call ----------
def adapt_for_shutterstock(api_key: str, base_title: str, base_desc: str, base_tags_csv: str,
                           max_tags: int, desc_max: int) -> Optional[Dict]:
    from openai import OpenAI
    client = OpenAI(api_key=api_key)

    system = (
        "You are preparing metadata for Shutterstock. "
        f"Return strict JSON with keys: title (<= {TITLE_MAX_CHARS} chars), "
        f"description (<= {desc_max} chars), tags (array up to {max_tags}). "
        "Rules: factual, concise, no brand names or private info for RF items, "
        "no keyword stuffing, tags are singular nouns/short phrases ordered by relevance."
    )
    user = (
        "Adapt the following base metadata for Shutterstock.\n\n"
        f"Base title: {base_title}\n"
        f"Base description: {base_desc}\n"
        f"Base tags (CSV): {base_tags_csv}\n\n"
        "Output JSON only."
    )

    resp = client.chat.completions.create(
        model=GPT_MODEL,
        temperature=0.2,
        messages=[
            {"role":"system","content":system},
            {"role":"user","content":user},
        ],
        response_format={"type":"json_object"},
    )
    text = resp.choices[0].message.content.strip()
    data = json.loads(text)

    title = (data.get("title") or "").strip()
    desc  = (data.get("description") or "").strip()
    tags  = [t.strip() for t in (data.get("tags") or []) if isinstance(t, str) and t.strip()]
    # normalize
    if len(title) > TITLE_MAX_CHARS:
        title = title[:TITLE_MAX_CHARS].rstrip(" ,.;:-")
    # de-dup and cap
    tags = list(dict.fromkeys(tags))[:max_tags]
    return {"title": title, "description": desc, "tags": tags}

# ---------- Main work ----------
def run(limit: Optional[int], force: bool, dry_run: bool):
    api_key = load_api_key()
    if not api_key:
        log.error("Missing OPENAI_API_KEY (config.json or env).")
        return 0

    wb, ws, headers, col = open_workbook()
    max_tags, desc_max = read_platform_rules(wb)
    log.info(f"SS limits → max_keywords={max_tags}, description_max_chars={desc_max}")

    # Ensure we can index the columns we care about
    need_cols = set(REQUIRED_COLS + ["file_name","last_seen"])
    missing = [c for c in need_cols if c not in col]
    if missing:
        log.error(f"Workbook missing required columns: {missing}")
        return 0

    updated = 0
    for r in range(2, ws.max_row+1):
        enabled = parse_bool(ws.cell(row=r, column=col["SS_enabled"]).value)
        if not enabled:
            continue

        # pull base
        btitle = (ws.cell(row=r, column=col["base_title"]).value or "").strip()
        bdesc  = (ws.cell(row=r, column=col["base_description"]).value or "").strip()
        btags  = (ws.cell(row=r, column=col["base_tags"]).value or "").strip()

        if not (btitle and bdesc and btags):
            # no base yet → cannot proceed
            cur_status = (ws.cell(row=r, column=col["SS_status"]).value or "").strip()
            if cur_status not in ("NOT_ENABLED","NEEDS_RELEASE","ERROR"):
                ws.cell(row=r, column=col["SS_status"], value="PENDING_PREP")
            continue

        # check current SS fields
        ss_title = (ws.cell(row=r, column=col["SS_title"]).value or "").strip()
        ss_desc  = (ws.cell(row=r, column=col["SS_description"]).value or "").strip()
        ss_tags  = (ws.cell(row=r, column=col["SS_tags"]).value or "").strip()

        already_filled = bool(ss_title and ss_desc and ss_tags)
        if already_filled and not force:
            continue  # leave manual/previous results intact

        # call GPT
        try:
            result = adapt_for_shutterstock(api_key, btitle, bdesc, btags, max_tags, desc_max)
        except Exception as e:
            log.exception(f"GPT call failed (row {r}): {e}")
            ws.cell(row=r, column=col["SS_status"], value="ERROR")
            continue

        title = result["title"]
        desc  = result["description"]
        tags  = result["tags"]

        ok, reason = small_qa(title, desc, tags, max_tags, desc_max)
        if not ok:
            log.warning(f"QA failed row {r}: {reason}")
            ws.cell(row=r, column=col["SS_status"], value="PENDING_PREP")
            # still write draft fields so you can fix by hand
            ws.cell(row=r, column=col["SS_title"],       value=title)
            ws.cell(row=r, column=col["SS_description"], value=desc)
            ws.cell(row=r, column=col["SS_tags"],        value=", ".join(tags))
        else:
            ws.cell(row=r, column=col["SS_title"],       value=title)
            ws.cell(row=r, column=col["SS_description"], value=desc)
            ws.cell(row=r, column=col["SS_tags"],        value=", ".join(tags))
            ws.cell(row=r, column=col["SS_status"],      value="READY_TO_UPLOAD")

        ws.cell(row=r, column=col["last_seen"], value=datetime.utcnow().isoformat())
        updated += 1

        if not dry_run:
            # Save after each row to avoid losing work on long runs
            wb.save(EXCEL_PATH)

        time.sleep(PER_ITEM_SLEEP)
        if limit and updated >= limit:
            break

    if not dry_run:
        wb.save(EXCEL_PATH)
    log.info(f"Shutterstock adaptation updated {updated} row(s)")
    return updated

# ---------- CLI ----------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Adapt base_* to SS_* (title/description/tags) for Shutterstock")
    parser.add_argument("--limit", type=int, default=None, help="Max rows to process")
    parser.add_argument("--force", action="store_true", help="Overwrite existing SS_* fields")
    parser.add_argument("--dry-run", action="store_true", help="Compute but do not write to disk")
    args = parser.parse_args()

    acquire_lock()
    try:
        n = run(limit=args.limit, force=args.force, dry_run=args.dry_run)
        print(f"✅ SS adapter updated {n} row(s)")
    except Exception as e:
        log.exception(f"Run failed: {e}")
        print("❌ Adaptation failed; see logs/adapt_ss.log")
    finally:
        _remove_lock()
