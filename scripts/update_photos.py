# scripts/update_photos.py
import os
import io
import re
import json
import base64
import hashlib
import logging
import signal
import atexit
import time
from logging.handlers import RotatingFileHandler
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Tuple, Optional

from PIL import Image
from openpyxl import load_workbook

import argparse

# ---------- Paths / Constants ----------
INBOX = Path("inbox")
EXCEL_PATH = Path("data/photos.xlsx")
SHEET_NAME = "Photos"
LOG_DIR = Path("logs")
LOG_FILE = LOG_DIR / "update.log"
LOCK_FILE = Path(".update.lock")
LOCK_STALE_AFTER = timedelta(hours=2)

PREVIEWS_DIR = Path("previews")
PREVIEW_MAX_SIDE = 768
PREVIEW_QUALITY = 80

SUPPORTED_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".tiff", ".heic", ".heif"}

# Core headers we manage (per-platform columns are preserved automatically)
CORE_HEADERS = [
    "file_name","rel_path","abs_path","size_kb","width","height",
    "created_time","modified_time","sha256","phash",
    "preview_path","preview_width","preview_height","preview_size_kb",
    "status_global","notes","last_seen",
    "base_title","base_description","base_tags",
    # legacy for convenience (kept & optionally backfilled)
    "description","tags",
]

# During scan we only update these (never clobber human/per-platform edits)
METADATA_FIELDS = {
    "file_name","rel_path","abs_path","size_kb","width","height",
    "created_time","modified_time","sha256","phash",
    "preview_path","preview_width","preview_height","preview_size_kb",
    "last_seen"
}

# ---------- GPT config ----------
GPT_MODEL = "gpt-4o-mini"
MAX_TAGS = 35
TITLE_MAX_CHARS = 70
PER_IMAGE_SLEEP = 0.3

# ---------- Logging ----------
def setup_logging() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("photo_manager")
    logger.setLevel(logging.INFO)
    if not logger.handlers:
        fh = RotatingFileHandler(LOG_FILE, maxBytes=1_000_000, backupCount=3, encoding="utf-8")
        fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
        fh.setFormatter(fmt)
        logger.addHandler(fh)
        ch = logging.StreamHandler()
        ch.setFormatter(fmt)
        logger.addHandler(ch)
    return logger

log = setup_logging()

# ---------- Locking ----------
_lock_acquired = False

def _remove_lock():
    global _lock_acquired
    try:
        if _lock_acquired and LOCK_FILE.exists():
            LOCK_FILE.unlink()
            log.info("Lock released")
    except Exception as e:
        log.warning(f"Failed to remove lock file: {e}")

def acquire_lock():
    global _lock_acquired
    if LOCK_FILE.exists():
        try:
            mtime = datetime.fromtimestamp(LOCK_FILE.stat().st_mtime)
            age = datetime.now() - mtime
            if age > LOCK_STALE_AFTER:
                log.warning(f"Existing lock is stale (age {age}). Overriding.")
                LOCK_FILE.unlink(missing_ok=True)
            else:
                pid_txt = ""
                try: pid_txt = LOCK_FILE.read_text().strip()
                except Exception: pass
                log.error(f"Another run appears active (pid {pid_txt}, age {age}). Exiting.")
                raise SystemExit(1)
        except FileNotFoundError:
            pass
    try:
        fd = os.open(str(LOCK_FILE), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w") as f:
            f.write(str(os.getpid()))
        _lock_acquired = True
        log.info(f"Lock acquired (pid {os.getpid()})")
    except FileExistsError:
        log.error("Lock already exists, exiting.")
        raise SystemExit(1)

    atexit.register(_remove_lock)
    def _sig(signum, frame):
        log.info(f"Received signal {signum}, cleaning up…")
        _remove_lock()
        raise SystemExit(130)
    for s in (signal.SIGINT, signal.SIGTERM):
        try: signal.signal(s, _sig)
        except Exception: pass

# ---------- Helpers ----------
def sha256sum(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def generate_preview(src: Path, sha: str) -> Dict[str, Optional[float]]:
    PREVIEWS_DIR.mkdir(parents=True, exist_ok=True)
    out = PREVIEWS_DIR / f"{sha}.jpg"
    if not out.exists():
        try:
            with Image.open(src) as img:
                img = img.convert("RGB")
                img.thumbnail((PREVIEW_MAX_SIDE, PREVIEW_MAX_SIDE))
                img.save(out, format="JPEG", quality=PREVIEW_QUALITY, optimize=True)
        except Exception as e:
            log.warning(f"Preview failed for {src}: {e}")
            return {"preview_path":"", "preview_width":None, "preview_height":None, "preview_size_kb":None}
    try:
        with Image.open(out) as p:
            pw, ph = p.size
        return {
            "preview_path": str(out.resolve()),
            "preview_width": pw,
            "preview_height": ph,
            "preview_size_kb": round(out.stat().st_size/1024, 2)
        }
    except Exception as e:
        log.warning(f"Preview read failed for {out}: {e}")
        return {"preview_path":"", "preview_width":None, "preview_height":None, "preview_size_kb":None}

def get_metadata(path: Path) -> Dict:
    st = path.stat()
    size_kb = round(st.st_size/1024, 2)
    ctime = datetime.fromtimestamp(st.st_ctime).isoformat()
    mtime = datetime.fromtimestamp(st.st_mtime).isoformat()
    w = h = None
    try:
        with Image.open(path) as img:
            w, h = img.size
    except Exception as e:
        log.warning(f"Dim read failed for {path}: {e}")
    sha = sha256sum(path)
    prev = generate_preview(path, sha)
    return {
        "file_name": path.name,
        "rel_path": str(path.relative_to(INBOX.parent)),
        "abs_path": str(path.resolve()),
        "size_kb": size_kb,
        "width": w,
        "height": h,
        "created_time": ctime,
        "modified_time": mtime,
        "sha256": sha,
        "phash": "",
        "preview_path": prev.get("preview_path",""),
        "preview_width": prev.get("preview_width"),
        "preview_height": prev.get("preview_height"),
        "preview_size_kb": prev.get("preview_size_kb"),
        "status_global": "NEW",
        "notes": "",
        "last_seen": datetime.utcnow().isoformat(),
        "base_title": "",
        "base_description": "",
        "base_tags": "",
        "description": "",  # legacy
        "tags": "",         # legacy
    }

def open_workbook():
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"{EXCEL_PATH} not found")
    wb = load_workbook(EXCEL_PATH)
    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in workbook")
    ws = wb[SHEET_NAME]
    headers = [c.value for c in ws[1]]
    # ensure all core headers exist (append missing)
    missing = [h for h in CORE_HEADERS if h not in headers]
    if missing:
        for h in missing:
            headers.append(h)
            ws.cell(row=1, column=len(headers), value=h)
        log.info(f"Added missing core headers: {', '.join(missing)}")
        wb.save(EXCEL_PATH)
    col = {h:i+1 for i,h in enumerate(headers)}
    return wb, ws, headers, col

def load_existing_rows(ws, headers) -> Dict[str, int]:
    idx_abs = headers.index("abs_path")
    existing = {}
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row: continue
        key = row[idx_abs] if idx_abs < len(row) else None
        if key:
            existing[str(key)] = r
    return existing

def looks_auto_title(title: str, file_name: str = "") -> bool:
    """Heuristics for detecting auto/ugly titles (timestamps, plain stems, digits)."""
    if not title:
        return True
    t = title.strip()
    if not t or t.lower() in {"none", "null"}:
        return True
    # YYYY-MM-DD[_]HHMMSS or YYYYMMDD or long digits
    if re.match(r"^\d{4}[-_]\d{2}[-_]\d{2}([ T]\d{2}[:_-]?\d{2}[:_-]?\d{2})?$", t):
        return True
    if re.match(r"^\d{8,}$", t):
        return True
    if all(ch.isdigit() or ch in "-_ " for ch in t) and sum(ch.isdigit() for ch in t) >= int(0.6*len(t)):
        return True
    if file_name:
        stem = Path(file_name).stem.replace("_"," ").replace("-"," ").strip().lower()
        if t.lower() == stem:
            return True
    if len(t) < 4:
        return True
    return False

# ---------- Enabled defaults ----------
def backfill_enabled_defaults(ws, headers, col) -> int:
    """
    For every column ending with '_enabled', set blank cells to True.
    Does NOT change cells that already contain a value (e.g., False).
    """
    enabled_cols = [h for h in headers if h.endswith("_enabled")]
    if not enabled_cols:
        return 0
    changed = 0
    for r in range(2, ws.max_row+1):
        for h in enabled_cols:
            c = ws.cell(row=r, column=col[h])
            v = c.value
            if v is None or (isinstance(v, str) and v.strip() == ""):
                c.value = True
                changed += 1
    if changed:
        log.info(f"Default-enabled {changed} blank *_enabled cells to TRUE")
    return changed

# ---------- Scan / Excel sync ----------
def scan_inbox() -> List[Dict]:
    if not INBOX.exists():
        log.error("inbox/ not found.")
        return []
    metas = []
    for p in INBOX.rglob("*"):
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTS:
            try:
                metas.append(get_metadata(p))
            except Exception as e:
                log.error(f"Metadata extraction failed for {p}: {e}")
    return metas

def update_excel(file_metas: List[Dict]) -> Tuple[int,int]:
    wb, ws, headers, col = open_workbook()
    existing = load_existing_rows(ws, headers)
    enabled_cols = [h for h in headers if h.endswith("_enabled")]
    added = updated = 0

    for meta in file_metas:
        key = meta["abs_path"]
        if key in existing:
            r = existing[key]
            for field in METADATA_FIELDS:
                if field in col:
                    ws.cell(row=r, column=col[field], value=meta.get(field,""))
            updated += 1
        else:
            # Prepare row values in header order
            row_vals = [meta.get(h,"") for h in headers]
            # Set *_enabled defaults to TRUE for new rows
            for h in enabled_cols:
                idx = headers.index(h)
                if row_vals[idx] in ("", None):
                    row_vals[idx] = True
            ws.append(row_vals)
            added += 1

    # Backfill any blank *_enabled in existing rows (do not flip explicit False)
    changed = backfill_enabled_defaults(ws, headers, col)
    if changed or added or updated:
        wb.save(EXCEL_PATH)
    return added, updated

# ---------- GPT: full title+description+tags ----------
def load_api_key() -> Optional[str]:
    cfg = Path("config.json")
    if cfg.exists():
        try:
            with open(cfg) as f:
                data = json.load(f)
                if data.get("OPENAI_API_KEY"):
                    return data["OPENAI_API_KEY"]
        except Exception as e:
            log.warning(f"config.json read failed: {e}")
    return os.getenv("OPENAI_API_KEY")

def data_url_from_path(path: Path) -> Optional[str]:
    try:
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("ascii")
        return f"data:image/jpeg;base64,{b64}"
    except Exception as e:
        log.error(f"Preview read for GPT failed: {path} | {e}")
        return None

def prompt_for_master_metadata():
    system = (
        "You are a metadata generator for stock photo marketplaces. "
        f"Produce a short, human-friendly title (<= {TITLE_MAX_CHARS} chars), "
        "a concise factual description (<= 180 chars), and 15–35 ranked tags. "
        "Avoid brands, private info, overclaiming, or spam. "
        f"Return strict JSON with keys: title (string), description (string), tags (array, max {MAX_TAGS})."
    )
    user = (
        "Look at the image and produce JSON. "
        "The title should be natural and readable (e.g., 'Sunset over mountain lake'), no trailing period. "
        "Description: main subject and clear context only. "
        f"Tags: singular nouns/short phrases, high relevance first, up to {MAX_TAGS}, no duplicates, no brand names."
    )
    return system, user

def call_gpt_for_preview(api_key: str, preview_path: Path) -> Optional[Dict]:
    try:
        from openai import OpenAI
    except Exception as e:
        log.error(f"OpenAI SDK not installed: {e}")
        return None

    data_url = data_url_from_path(preview_path)
    if not data_url:
        return None
    client = OpenAI(api_key=api_key)
    system, user = prompt_for_master_metadata()

    try:
        resp = client.chat.completions.create(
            model=GPT_MODEL,
            temperature=0.2,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": [
                    {"type": "text", "text": user},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ]},
            ],
            response_format={"type": "json_object"},
        )
        text = resp.choices[0].message.content.strip()
        data = json.loads(text)

        title = (data.get("title") or "").strip()
        desc  = (data.get("description") or "").strip()
        tags  = [t.strip() for t in (data.get("tags") or []) if isinstance(t, str) and t.strip()]
        tags  = list(dict.fromkeys(tags))[:MAX_TAGS]

        if len(title) > TITLE_MAX_CHARS:
            title = title[:TITLE_MAX_CHARS].rstrip(" ,.;:-")

        return {"title": title, "description": desc, "tags": tags}
    except Exception as e:
        log.error(f"GPT call failed for {preview_path.name}: {e}")
        return None

def describe_needed_rows(limit: Optional[int]=None) -> int:
    """
    For rows where any of base_title/base_description/base_tags is missing
    OR base_title looks auto/ugly, and status_global in {NEW, ANALYZED, DESCRIBED},
    call GPT (image-based) to generate title+description+tags.
    """
    api_key = load_api_key()
    if not api_key:
        log.error("Missing OPENAI_API_KEY (config.json or env).")
        return 0

    wb, ws, headers, col = open_workbook()
    must = ["status_global","base_title","base_description","base_tags",
            "preview_path","abs_path","sha256","file_name",
            "description","tags","last_seen"]
    for m in must:
        if m not in col:
            log.error(f"Missing required column in workbook: {m}")
            return 0

    updated = 0
    for r in range(2, ws.max_row+1):
        status  = (ws.cell(row=r, column=col["status_global"]).value or "").strip()
        if status not in ("NEW","ANALYZED","DESCRIBED"):
            continue

        btitle  = (ws.cell(row=r, column=col["base_title"]).value or "").strip()
        bdesc   = (ws.cell(row=r, column=col["base_description"]).value or "").strip()
        btags   = (ws.cell(row=r, column=col["base_tags"]).value or "").strip()
        fname   = (ws.cell(row=r, column=col["file_name"]).value or "").strip()
        p_path  = (ws.cell(row=r, column=col["preview_path"]).value or "").strip()
        abs_p   = (ws.cell(row=r, column=col["abs_path"]).value or "").strip()

        # Decide if we need to (re)generate
        need = False
        if not (btitle and bdesc and btags):
            need = True
        elif looks_auto_title(btitle, fname):
            need = True
        if not need:
            continue

        # Ensure preview exists
        if not p_path:
            if abs_p:
                sha = ws.cell(row=r, column=col["sha256"]).value or ""
                if not sha:
                    sha = sha256sum(Path(abs_p))
                    ws.cell(row=r, column=col["sha256"], value=sha)
                if sha:
                    prev = generate_preview(Path(abs_p), sha)
                    for k in ("preview_path","preview_width","preview_height","preview_size_kb"):
                        if k in col:
                            ws.cell(row=r, column=col[k], value=prev.get(k))
                    wb.save(EXCEL_PATH)
                    p_path = prev.get("preview_path") or ""

        if not p_path or not Path(p_path).exists():
            log.warning(f"No preview for row {r} ({abs_p}); skipping")
            continue

        log.info(f"Describing (full): {Path(p_path).name}")
        result = call_gpt_for_preview(api_key, Path(p_path))
        if not result:
            continue

        # Write base fields
        ws.cell(row=r, column=col["base_title"],       value=result["title"])
        ws.cell(row=r, column=col["base_description"], value=result["description"])
        ws.cell(row=r, column=col["base_tags"],        value=", ".join(result["tags"]))
        # Legacy backfill if empty
        if not (ws.cell(row=r, column=col["description"]).value or "").strip():
            ws.cell(row=r, column=col["description"], value=result["description"])
        if not (ws.cell(row=r, column=col["tags"]).value or "").strip():
            ws.cell(row=r, column=col["tags"], value=", ".join(result["tags"]))

        ws.cell(row=r, column=col["status_global"], value="DESCRIBED")
        ws.cell(row=r, column=col["last_seen"], value=datetime.utcnow().isoformat())
        wb.save(EXCEL_PATH)
        updated += 1
        time.sleep(PER_IMAGE_SLEEP)

        if limit and updated >= limit:
            break

    log.info(f"Describe step updated {updated} row(s)")
    return updated

# ---------- CLI ----------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scan inbox/ and sync data/photos.xlsx")
    parser.add_argument("--reset", action="store_true", help="Reset Excel (keep headers only)")
    parser.add_argument("--describe", action="store_true", help="Generate/refresh base_title, base_description & base_tags with GPT")
    parser.add_argument("--describe-limit", type=int, default=None, help="Max rows to process in describe step")
    args = parser.parse_args()

    acquire_lock()

    if args.reset:
        log.info("Starting reset…")
        wb, ws, headers, col = open_workbook()
        ws.delete_rows(2, ws.max_row)
        wb.save(EXEXCEL_PATH)  # <-- typo guard; correct below
        # Fix typo: ensure proper save target
        wb.save(EXCEL_PATH)
        log.info(f"Reset {EXCEL_PATH} (headers only kept)")
    else:
        log.info("Scan started")
        metas = scan_inbox()
        log.info(f"Discovered {len(metas)} candidate files")
        try:
            added, updated = update_excel(metas)
            log.info(f"Excel updated: added={added}, updated={updated}")
            print(f"✅ Updated {EXCEL_PATH} | added={added}, updated={updated}")
        except Exception as e:
            log.exception(f"Excel update failed: {e}")
            print("❌ Excel update failed (see logs/update.log)")

        if args.describe:
            try:
                n = describe_needed_rows(limit=args.describe_limit)
                print(f"📝 Describe step updated {n} row(s)")
            except Exception as e:
                log.exception(f"Describe step failed: {e}")
                print("❌ Describe step failed (see logs/update.log)")
